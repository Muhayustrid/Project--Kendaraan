from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl ,xlrd
from openpyxl import Workbook
import pathlib

root=Tk()
root.title("Data Input")
root.geometry('700x400+300+200')
root.resizable(False,False)
root.configure(bg="#326273")

file=pathlib.Path('tes1.xlsx')
if file.exists():
    pass
    file=Workbook()
    sheet=file.active
    sheet['A1']="jenis kendaraan"
    sheet['B1']="Merk"
    sheet['C1']="Tahun Produksi"
    sheet['D1']="warna"
    sheet['E1']="Nomor Polisi"
    sheet['F1']="Nama Pemilik"
    file.save('tes1.xlsx')

else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="jenis kendaraan"
    sheet['B1']="Merk"
    sheet['C1']="Tahun Produksi"
    sheet['D1']="warna"
    sheet['E1']="Nomor Polisi"
    sheet['F1']="Nama Pemilik"
    file.save('tes1.xlsx')

def submit():
    jeniskendaraan=jeniskendaraan_combobox.get()
    merk=merk_combobox.get()
    age=age_combobox.get()
    warna=warna_combobox.get()
    namapemilik=namapemilikEntry.get()
    Nopol=NopolEntry.get(1.0,END)
    

    file=openpyxl.load_workbook('tes1.xlsx')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=jeniskendaraan)
    sheet.cell(column=2,row=sheet.max_row,value=merk)
    sheet.cell(column=3,row=sheet.max_row,value=age)
    sheet.cell(column=4,row=sheet.max_row,value=warna)
    sheet.cell(column=5,row=sheet.max_row,value=Nopol)
    sheet.cell(column=6,row=sheet.max_row,value=namapemilik)
    
    file.save(r'tes1.xlsx')

    messagebox.showinfo('Notifikasi','INPO DITERIMA!')

    #jeniskendaraanValue.set('')
    #merkValue.set('')
    #ageValue.set('')
    NopolEntry.delete(1.0,END)

def clear():
    namapemilikValue.set('')
    jeniskendaraan_combobox.set('')
    merk_combobox.set('')
    age_combobox.set('')
    warna_combobox.set('')
    NopolEntry.delete(1.0,END)

#Untuk Icon
icon_image=PhotoImage(file="Icon.png")
root.iconphoto(False,icon_image)

#Untuk Header
Label(root,text="Silahkan Isi kolom-kolom tersebut:",font="arial 13", bg="#326273",fg="#fff").place(x=20,y=20)

#Label
Label(root,text='Jenis Kendaraan',font=19,bg='#326273',fg="#fff").place(x=50,y=100)
Label(root,text='Merk Kendaraan',font=19,bg='#326273',fg="#fff").place(x=50,y=150)
Label(root,text='Tahun Produksi',font=19,bg='#326273',fg="#fff").place(x=50,y=200)
Label(root,text='Warna',font=19,bg='#326273',fg="#fff").place(x=50,y=300)
Label(root,text='Nomor Polisi',font=19,bg='#326273',fg="#fff").place(x=50,y=250)
Label(root,text='Nama Pemilik',font=19,bg='#326273',fg="#fff").place(x=310,y=100)

#Entry
namapemilikValue = StringVar()

namapemilikEntry = Entry(root,textvariable=namapemilikValue,width=15,bd=2,font=20)

#Jenis Kendaraan
jeniskendaraan_combobox = Combobox(root,values=['Mobil', 'Motor', 'Truk', 'Bus', 'Mobil Pick-up', 'Mobil Box', 'Mobil Besar', 'Mobil Kecil', 'Mobil Sedan', 'Mobil SUV', 'Mobil MPV', 'Mobil Coupé', 'Mobil Convertible', 'Mobil Sport', 'Mobil Ekskavator', 'Mobil Bulldozer', 'Mobil Grader', 'Mobil Paver', 'Mobil Crane', 'Mobil Dump Truck', 'Mobil Tanker', 'Mobil Trailer', 'Mobil Alat Berat Lainnya'],font='arial 9',state='r',width=11)
jeniskendaraan_combobox.set('')
jeniskendaraan_combobox.place(x=200,y=105)

#Merk Kendaraan
merk_combobox = Combobox(root,values=['Yamaha','Suzuki', 'Toyota', 'Honda', 'Mitsubishi', 'Nissan', 'Mazda', 'Daihatsu', 'Hyundai', 'Kia', 'Ford', 'Chevrolet', 'Mercedes-Benz', 'BMW', 'Audi', 'Volkswagen', 'Peugeot', 'Renault', 'Citroen', 'Opel', 'Fiat', 'Skoda', 'Seat', 'Isuzu', 'Merkur', 'Porsche', 'Lamborghini', 'Ferarri', 'Bugatti', 'Bentley', 'Rolls-Royce']
,font='arial 9',state='r',width=11)
merk_combobox.set('')
merk_combobox.place(x=200,y=155)

#Tahun Produksi
age_combobox = Combobox(root,values=['2022', '2021', '2020', '2019', '2018', '2017', '2016', '2015', '2014', '2013', '2012', '2011', '2010', '2009', '2008', '2007', '2006', '2005', '2004', '2003', '2002', '2001', '2000', '1999', '1998', '1997', '1996', '1995', '1994', '1993', '1992', '1991', '1990', '1989', '1988', '1987', '1986', '1985', '1984', '1983', '1982', '1981', '1980', '1979', '1978', '1977', '1976', '1975', '1974', '1973', '1972', '1971', '1970']
,font='arial 9',state='r',width=11)
age_combobox.set('')
age_combobox.place(x=200,y=205)

#TahunProduksi
warna_combobox = Combobox(root,values=['Merah', 'Kuning', 'Biru', 'Hitam', 'Putih', 'Abu-abu', 'Hijau', 'Jingga', 'Ungu', 'Coklat', 'Biru Muda', 'Hijau Muda', 'Biru Tua', 'Merah Muda', 'Merah Tua', 'Kuning Muda', 'Kuning Tua', 'Ungu Muda', 'Ungu Tua', 'Pink', 'Orange', 'Cokelat Tua', 'Cokelat Muda', 'Abu-abu Muda', 'Abu-abu Tua', 'Hijau Tua', 'Hijau Muda', 'Jingga Muda', 'Jingga Tua'],font='arial 9',state='r',width=11)
warna_combobox.set('')
warna_combobox.place(x=200,y=300)

NopolEntry = Text(root,width=11,height=1,bd=2)

Button(root,text="Kirim",bg="#326273",fg="white",width=15,height=2,command=submit).place(x=200,y=350)
Button(root,text="Clear",bg="#326273",fg="white",width=15,height=2,command=clear).place(x=340,y=350)
Button(root,text="Keluar",bg="#326273",fg="white",width=15,height=2,command=lambda:root.destroy()).place(x=480,y=350)


namapemilikEntry.place(x=450,y=100)
NopolEntry.place(x=200,y=250)
root.mainloop()
